import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import FormulaRule


def set_col_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def add_table(ws, ref, name):
    table = Table(displayName=name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)


def add_header(ws, headers):
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill
    ws.freeze_panes = "A2"


def build_workbook():
    wb = Workbook()

    # Lookups sheet
    ws_lookup = wb.active
    ws_lookup.title = "Lookups"
    add_header(ws_lookup, [
        "Trade", "Issue_Type", "Status", "Relation", "Handover_Type", "Resource_Type"
    ])
    lookups_rows = [
        ["GC", "设计", "Open", "FS", "结构", "吊篮"],
        ["幕墙", "材料", "In-Progress", "SS", "砌筑", "塔吊"],
        ["门窗", "场地", "Closed", "FF", "抹灰", "气吊"],
        ["机电", "移交", None, "SF", "屋面", "班组"],
        ["精装", "验收", None, None, "立面", None],
        [None, "天气", None, None, None, None],
    ]
    for r in lookups_rows:
        ws_lookup.append(r)
    set_col_widths(ws_lookup, [16, 16, 16, 12, 18, 16])

    # Data Validations referencing Lookups
    trade_range = "$A$2:$A$100"
    issue_type_range = "$B$2:$B$100"
    status_range = "$C$2:$C$100"
    relation_range = "$D$2:$D$100"
    handover_type_range = "$E$2:$E$100"
    resource_type_range = "$F$2:$F$100"

    # Master_Plan
    ws_plan = wb.create_sheet("Master_Plan")
    plan_headers = [
        "WBS_ID", "Node_Name", "Building", "Floor", "Trade", "Planned_Start",
        "Planned_Finish", "Duration_d", "Crew_Size", "Preconditions", "Accept_Criteria",
        "Milestone_Flag", "Total_Float_d", "RAG"
    ]
    add_header(ws_plan, plan_headers)
    set_col_widths(ws_plan, [22, 26, 12, 10, 12, 16, 16, 12, 12, 28, 26, 14, 14, 10])

    # Sample plan data based on the facade example
    # Dates use 2025-08 for illustration
    sample_plan = [
        ["B01-F-EXT-001", "外墙抹灰完成", "B01", "ALL", "GC", datetime.date(2025, 8, 1), datetime.date(2025, 8, 5), 5, 12, "", "检验批合格", "N", 1, "Yellow"],
        ["B01-F-EXT-002", "外墙第一遍腻子", "B01", "ALL", "GC", datetime.date(2025, 8, 6), datetime.date(2025, 8, 8), 3, 10, "", "检验批合格", "N", 1, "Yellow"],
        ["B01-F-EXT-003", "拆外架", "B01", "ALL", "GC", datetime.date(2025, 8, 9), datetime.date(2025, 8, 10), 2, 8, "外墙抹灰+一遍腻子完成", "安全验收通过", "N", 1, "Yellow"],
        ["B01-RF-001", "屋面保护层浇筑完成", "B01", "RF", "GC", datetime.date(2025, 8, 1), datetime.date(2025, 8, 5), 5, 8, "", "检验批合格", "N", 5, "Green"],
        ["B01-RF-002", "屋面保护层养护达标(≥10d)", "B01", "RF", "GC", datetime.date(2025, 8, 5), datetime.date(2025, 8, 15), 11, 0, "屋面保护层完成", "养护记录合格", "Y", 3, "Green"],
        ["B01-GONDL-001", "吊篮安装", "B01", "ALL", "幕墙", datetime.date(2025, 8, 16), datetime.date(2025, 8, 16), 1, 6, "外架拆除完成+屋面养护≥10天", "设备验收合格", "N", 2, "Green"],
        ["B01-PAINT-001", "外立面涂料开始", "B01", "ALL", "GC", datetime.date(2025, 8, 17), datetime.date(2025, 8, 25), 9, 10, "吊篮到位", "抽检合格", "N", 0, "Yellow"],
        ["B01-GLASS-001", "玻璃/石材/铝板安装开始", "B01", "ALL", "幕墙", datetime.date(2025, 8, 17), datetime.date(2025, 8, 30), 14, 12, "吊篮到位", "抽检合格", "N", 0, "Yellow"],
        ["B01-SEAL-001", "打胶密封", "B01", "ALL", "幕墙", datetime.date(2025, 8, 31), datetime.date(2025, 9, 3), 4, 6, "立面安装完成", "淋水试验合格", "N", 0, "Yellow"],
        ["B01-HANDOVER-001", "成品保护与清理+分段验收", "B01", "ALL", "幕墙", datetime.date(2025, 9, 4), datetime.date(2025, 9, 6), 3, 6, "", "验收合格", "Y", 0, "Yellow"],
    ]
    for r in sample_plan:
        ws_plan.append(r)

    # Data validation for Trade (E), Milestone_Flag (L), RAG (N)
    dv_trade = DataValidation(type="list", formula1=f"=Lookups!{trade_range}", allow_blank=True)
    dv_milestone = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    dv_rag = DataValidation(type="list", formula1='"Green,Yellow,Red"', allow_blank=True)
    ws_plan.add_data_validation(dv_trade)
    ws_plan.add_data_validation(dv_milestone)
    ws_plan.add_data_validation(dv_rag)
    dv_trade.add(f"E2:E2000")
    dv_milestone.add(f"L2:L2000")
    dv_rag.add(f"N2:N2000")

    # Conditional formatting on RAG text
    ws_plan.conditional_formatting.add(
        "N2:N2000",
        FormulaRule(formula=["$N2=\"Red\""], fill=PatternFill("solid", fgColor="FFC7CE"))
    )
    ws_plan.conditional_formatting.add(
        "N2:N2000",
        FormulaRule(formula=["$N2=\"Yellow\""], fill=PatternFill("solid", fgColor="FFEB9C"))
    )
    ws_plan.conditional_formatting.add(
        "N2:N2000",
        FormulaRule(formula=["$N2=\"Green\""], fill=PatternFill("solid", fgColor="C6EFCE"))
    )

    # Turn Master_Plan into a table
    last_row_plan = ws_plan.max_row
    add_table(ws_plan, f"A1:N{last_row_plan}", "tblMasterPlan")

    # Actuals
    ws_act = wb.create_sheet("Actuals")
    act_headers = [
        "WBS_ID", "Report_Date", "Actual_Start", "Actual_Finish", "%Complete",
        "Daily_Quantity", "Cum_Quantity", "Blockers", "PIC"
    ]
    add_header(ws_act, act_headers)
    set_col_widths(ws_act, [22, 14, 16, 16, 12, 14, 14, 24, 14])

    sample_actuals = [
        ["B01-F-EXT-001", datetime.date(2025, 8, 7), datetime.date(2025, 8, 1), datetime.date(2025, 8, 7), 100, None, None, None, "张三"],
        ["B01-F-EXT-002", datetime.date(2025, 8, 11), datetime.date(2025, 8, 9), datetime.date(2025, 8, 11), 100, None, None, None, "张三"],
        ["B01-F-EXT-003", datetime.date(2025, 8, 12), datetime.date(2025, 8, 11), datetime.date(2025, 8, 12), 100, None, None, None, "李四"],
        ["B01-RF-001", datetime.date(2025, 8, 5), datetime.date(2025, 8, 1), datetime.date(2025, 8, 5), 100, None, None, None, "王五"],
        ["B01-RF-002", datetime.date(2025, 8, 15), datetime.date(2025, 8, 5), datetime.date(2025, 8, 15), 100, None, None, None, "王五"],
        ["B01-GONDL-001", datetime.date(2025, 8, 16), datetime.date(2025, 8, 16), None, 5, None, None, None, "赵六"],
    ]
    for r in sample_actuals:
        ws_act.append(r)
    add_table(ws_act, f"A1:I{ws_act.max_row}", "tblActuals")

    # Dependencies
    ws_dep = wb.create_sheet("Dependencies")
    dep_headers = ["Pre_WBS_ID", "Succ_WBS_ID", "Relation", "Lag_d"]
    add_header(ws_dep, dep_headers)
    set_col_widths(ws_dep, [22, 22, 12, 10])

    dep_samples = [
        ["B01-F-EXT-002", "B01-F-EXT-003", "FS", 0],
        ["B01-RF-002", "B01-GONDL-001", "FS", 0],  # 养护达标后可装吊篮
        ["B01-F-EXT-003", "B01-GONDL-001", "FS", 0],  # 无外架+屋面养护并行满足
        ["B01-GONDL-001", "B01-PAINT-001", "FS", 0],
        ["B01-GONDL-001", "B01-GLASS-001", "FS", 0],
        ["B01-GLASS-001", "B01-SEAL-001", "FS", 0],
        ["B01-SEAL-001", "B01-HANDOVER-001", "FS", 0],
    ]
    for r in dep_samples:
        ws_dep.append(r)
    dv_relation = DataValidation(type="list", formula1=f"=Lookups!{relation_range}", allow_blank=True)
    ws_dep.add_data_validation(dv_relation)
    dv_relation.add("C2:C2000")
    add_table(ws_dep, f"A1:D{ws_dep.max_row}", "tblDependencies")

    # Thresholds
    ws_thr = wb.create_sheet("Thresholds")
    thr_headers = ["Trade", "RAG_SPI_Green", "RAG_SPI_Yellow", "Float_Green_d", "Float_Yellow_d", "Lookahead_Weeks"]
    add_header(ws_thr, thr_headers)
    set_col_widths(ws_thr, [16, 16, 18, 16, 18, 18])

    thr_rows = [
        ["GC", 0.95, 0.85, 3, 2, 4],
        ["幕墙", 0.95, 0.85, 3, 2, 4],
        ["门窗", 0.95, 0.85, 3, 2, 4],
        ["机电", 0.95, 0.85, 3, 2, 4],
        ["精装", 0.95, 0.85, 3, 2, 4],
    ]
    for r in thr_rows:
        ws_thr.append(r)
    dv_thr_trade = DataValidation(type="list", formula1=f"=Lookups!{trade_range}", allow_blank=True)
    ws_thr.add_data_validation(dv_thr_trade)
    dv_thr_trade.add("A2:A2000")
    add_table(ws_thr, f"A1:F{ws_thr.max_row}", "tblThresholds")

    # Issues_Log
    ws_iss = wb.create_sheet("Issues_Log")
    iss_headers = ["Issue_ID", "Date", "WBS_ID", "Building", "Issue_Type", "Description", "Owner", "Due_Date", "Status", "Attachment"]
    add_header(ws_iss, iss_headers)
    set_col_widths(ws_iss, [16, 14, 22, 12, 14, 36, 14, 14, 16, 26])

    iss_rows = [
        ["ISS-0001", datetime.date(2025, 8, 10), "B01-F-EXT-003", "B01", "场地", "外架拆除工序与内装交叉，需夜间施工", "总包-现场经理", datetime.date(2025, 8, 12), "In-Progress", None]
    ]
    for r in iss_rows:
        ws_iss.append(r)
    dv_issue_type = DataValidation(type="list", formula1=f"=Lookups!{issue_type_range}", allow_blank=True)
    dv_issue_status = DataValidation(type="list", formula1=f"=Lookups!{status_range}", allow_blank=True)
    ws_iss.add_data_validation(dv_issue_type)
    ws_iss.add_data_validation(dv_issue_status)
    dv_issue_type.add("E2:E2000")
    dv_issue_status.add("I2:I2000")
    add_table(ws_iss, f"A1:J{ws_iss.max_row}", "tblIssues")

    # Handovers
    ws_hov = wb.create_sheet("Handovers")
    hov_headers = ["From_Trade", "To_Trade", "Building", "Area/Axis", "Handover_Type", "Planned_Date", "Actual_Date", "Accept_Result"]
    add_header(ws_hov, hov_headers)
    set_col_widths(ws_hov, [16, 16, 12, 16, 18, 16, 16, 16])

    hov_rows = [
        ["GC", "幕墙", "B01", "A~C/1~4轴", "立面", datetime.date(2025, 8, 16), datetime.date(2025, 8, 16), "合格"]
    ]
    for r in hov_rows:
        ws_hov.append(r)
    dv_from_trade = DataValidation(type="list", formula1=f"=Lookups!{trade_range}", allow_blank=True)
    dv_to_trade = DataValidation(type="list", formula1=f"=Lookups!{trade_range}", allow_blank=True)
    dv_handover_type = DataValidation(type="list", formula1=f"=Lookups!{handover_type_range}", allow_blank=True)
    ws_hov.add_data_validation(dv_from_trade)
    ws_hov.add_data_validation(dv_to_trade)
    ws_hov.add_data_validation(dv_handover_type)
    dv_from_trade.add("A2:A2000")
    dv_to_trade.add("B2:B2000")
    dv_handover_type.add("E2:E2000")
    add_table(ws_hov, f"A1:H{ws_hov.max_row}", "tblHandovers")

    # Resources
    ws_res = wb.create_sheet("Resources")
    res_headers = ["Resource_Type", "Qty", "Location", "Booking_Slot", "Owner", "Conflicts"]
    add_header(ws_res, res_headers)
    set_col_widths(ws_res, [18, 10, 18, 20, 16, 24])

    res_rows = [
        ["吊篮", 6, "B01 东立面", "2025-08-16 ~ 2025-09-10", "幕墙-设备主管", "与B02冲突(建议错峰)"]
    ]
    for r in res_rows:
        ws_res.append(r)
    dv_resource_type = DataValidation(type="list", formula1=f"=Lookups!{resource_type_range}", allow_blank=True)
    ws_res.add_data_validation(dv_resource_type)
    dv_resource_type.add("A2:A2000")
    add_table(ws_res, f"A1:F{ws_res.max_row}", "tblResources")

    # Dashboard
    ws_dash = wb.create_sheet("Dashboard")
    ws_dash["A1"].value = "Week_Start"
    ws_dash["B1"].value = datetime.date(2025, 8, 11)  # Monday
    ws_dash["A2"].value = "Week_End"
    ws_dash["B2"].value = datetime.date(2025, 8, 17)  # Sunday
    ws_dash["A4"].value = "PCR (Project)"
    ws_dash["B4"].value = "=IFERROR(COUNTIFS(Master_Plan!G:G,\">=""&B1,Master_Plan!G:G, ""<=""&B2)/COUNTA(Master_Plan!A:A),0)"
    ws_dash["A5"].value = "ACR (Project)"
    ws_dash["B5"].value = "=IFERROR(COUNTIFS(Actuals!D:D,\">=""&B1,Actuals!D:D, ""<=""&B2)/COUNTA(Master_Plan!A:A),0)"
    ws_dash["A6"].value = "SPI (Project)"
    ws_dash["B6"].value = "=IFERROR(B5/B4,0)"

    # Trade-level table headers
    ws_dash["A9"].value = "Trade"
    ws_dash["B9"].value = "PCR"
    ws_dash["C9"].value = "ACR"
    ws_dash["D9"].value = "SPI"

    # Fill trades from Lookups (unique first 5 non-empty)
    trades = ["GC", "幕墙", "门窗", "机电", "精装"]
    start_row = 10
    for idx, t in enumerate(trades):
        r = start_row + idx
        ws_dash[f"A{r}"].value = t
        # PCR by trade
        ws_dash[f"B{r}"].value = f"=IFERROR(COUNTIFS(Master_Plan!E:E,A{r},Master_Plan!G:G,\">=""&$B$1,Master_Plan!G:G, ""<=""&$B$2)/COUNTIF(Master_Plan!E:E,A{r}),0)"
        # ACR by trade
        ws_dash[f"C{r}"].value = f"=IFERROR(COUNTIFS(Actuals!A:A,Master_Plan!A:A,Actuals!D:D,\">=""&$B$1),0)"  # placeholder, replaced below
        # For ACR by trade, count WBS in Actuals where its trade equals A{r}
        # Using SUMPRODUCT across Master_Plan trade matching and Actuals finish in window
        ws_dash[f"C{r}"].value = f"=IFERROR(SUMPRODUCT((Master_Plan!E:E=A{r})*(ISNUMBER(MATCH(Master_Plan!A:A,Actuals!A:A,0)))*(INDEX(Actuals!D:D,MATCH(Master_Plan!A:A,Actuals!A:A,0))>=$B$1)*(INDEX(Actuals!D:D,MATCH(Master_Plan!A:A,Actuals!A:A,0))<=$B$2))/COUNTIF(Master_Plan!E:E,A{r}),0)"
        ws_dash[f"D{r}"].value = f"=IFERROR(C{r}/B{r},0)"

    # KPI formatting
    ws_dash["A4"].font = Font(bold=True)
    ws_dash["A5"].font = Font(bold=True)
    ws_dash["A6"].font = Font(bold=True)
    ws_dash["B4"].number_format = "0.00%"
    ws_dash["B5"].number_format = "0.00%"
    ws_dash["B6"].number_format = "0.00"

    # Add a simple RAG indicator cell for project SPI vs default thresholds (from GC row)
    ws_dash["A8"].value = "Project RAG (by SPI)"
    ws_dash["B8"].value = "=IF(B6>=VLOOKUP(\"GC\",Thresholds!A:F,2,FALSE),\"Green\",IF(B6>=VLOOKUP(\"GC\",Thresholds!A:F,3,FALSE),\"Yellow\",\"Red\"))"

    # Notes/Instructions sheet
    ws_readme = wb.create_sheet("README")
    ws_readme["A1"].value = (
        "使用说明：\n"
        "1) 先在 Lookups 维护下拉枚举（如 Trade/Issue_Type 等）。\n"
        "2) 在 Master_Plan 维护基线计划；Total_Float_d 可由CPM计算后回填。\n"
        "3) 在 Actuals 每日填报，最终完成时填写 Actual_Finish。\n"
        "4) Dependencies 维护工序逻辑与时差（Lag_d）。\n"
        "5) Thresholds 可按专业设置 RAG 阈值。\n"
        "6) Dashboard 输入周起止日，自动计算 PCR/ACR/SPI（项目/专业）。\n"
        "7) Master_Plan 的 RAG 目前按文本标记；后续可由计算层自动写回。\n"
    )
    ws_readme.merge_cells("A1:F12")
    ws_readme["A1"].alignment = Alignment(wrap_text=True, vertical="top")

    # Styling for date columns
    date_cols_plan = [6, 7]
    for row in ws_plan.iter_rows(min_row=2, max_row=ws_plan.max_row, min_col=1, max_col=len(plan_headers)):
        for idx in date_cols_plan:
            row[idx-1].number_format = "yyyy-mm-dd"

    for row in ws_act.iter_rows(min_row=2, max_row=ws_act.max_row, min_col=1, max_col=ws_act.max_column):
        for idx in [2, 3, 4]:
            row[idx-1].number_format = "yyyy-mm-dd"

    for row in ws_hov.iter_rows(min_row=2, max_row=ws_hov.max_row, min_col=1, max_col=ws_hov.max_column):
        for idx in [6, 7]:
            row[idx-1].number_format = "yyyy-mm-dd"

    # Save workbook
    return wb


if __name__ == "__main__":
    wb = build_workbook()
    output_path = "/workspace/施工进度管控模板_v1.xlsx"
    wb.save(output_path)
    print(f"Saved: {output_path}")